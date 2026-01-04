"""
Excel Builder Service
Creates Excel files with VLOOKUP, HLOOKUP, formulas, and multiple sheets with graphs
"""
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any, Optional
import json


class ExcelBuilderService:
    """Service to build Excel files with advanced features"""
    
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Sheet1"
        self.sheets = {"Sheet1": self.ws}
        self.formulas = []
        self.charts = []
    
    def create_sheet(self, name: str, make_active: bool = False):
        """Create a new sheet"""
        if name in self.sheets:
            return self.sheets[name]
        
        ws = self.wb.create_sheet(title=name)
        self.sheets[name] = ws
        if make_active:
            self.wb.active = ws
        return ws
    
    def add_data(self, sheet_name: str, data: List[List[Any]], start_row: int = 1, start_col: int = 1):
        """Add data to a sheet"""
        ws = self.sheets.get(sheet_name, self.ws)
        
        for row_idx, row_data in enumerate(data, start=start_row):
            for col_idx, value in enumerate(row_data, start=start_col):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
        
        return ws
    
    def add_vlookup(self, 
                   sheet_name: str, 
                   target_cell: str,
                   lookup_value: str,
                   table_array: str,
                   col_index_num: int,
                   range_lookup: bool = False,
                   lookup_sheet: Optional[str] = None):
        """
        Add VLOOKUP formula
        
        Args:
            sheet_name: Target sheet
            target_cell: Cell to place formula (e.g., "B2")
            lookup_value: Value to search for (cell reference or value)
            table_array: Range to search (e.g., "A1:D10")
            col_index_num: Column number in table_array (1-based)
            range_lookup: False for exact match, True for approximate
            lookup_sheet: Sheet name for table_array (if different)
        """
        ws = self.sheets.get(sheet_name, self.ws)
        
        # Build table array reference
        if lookup_sheet:
            table_ref = f"'{lookup_sheet}'!{table_array}"
        else:
            table_ref = table_array
        
        # Create VLOOKUP formula
        formula = f'=VLOOKUP({lookup_value},{table_ref},{col_index_num},{str(range_lookup).upper()})'
        
        ws[target_cell] = formula
        
        self.formulas.append({
            'type': 'VLOOKUP',
            'cell': target_cell,
            'formula': formula,
            'sheet': sheet_name
        })
        
        return formula
    
    def add_hlookup(self,
                   sheet_name: str,
                   target_cell: str,
                   lookup_value: str,
                   table_array: str,
                   row_index_num: int,
                   range_lookup: bool = False,
                   lookup_sheet: Optional[str] = None):
        """
        Add HLOOKUP formula
        
        Args:
            sheet_name: Target sheet
            target_cell: Cell to place formula (e.g., "B2")
            lookup_value: Value to search for
            table_array: Range to search (e.g., "A1:D10")
            row_index_num: Row number in table_array (1-based)
            range_lookup: False for exact match, True for approximate
            lookup_sheet: Sheet name for table_array (if different)
        """
        ws = self.sheets.get(sheet_name, self.ws)
        
        if lookup_sheet:
            table_ref = f"'{lookup_sheet}'!{table_array}"
        else:
            table_ref = table_array
        
        formula = f'=HLOOKUP({lookup_value},{table_ref},{row_index_num},{str(range_lookup).upper()})'
        
        ws[target_cell] = formula
        
        self.formulas.append({
            'type': 'HLOOKUP',
            'cell': target_cell,
            'formula': formula,
            'sheet': sheet_name
        })
        
        return formula
    
    def add_formula(self, sheet_name: str, cell: str, formula: str):
        """Add any Excel formula"""
        ws = self.sheets.get(sheet_name, self.ws)
        ws[cell] = formula
        
        self.formulas.append({
            'type': 'CUSTOM',
            'cell': cell,
            'formula': formula,
            'sheet': sheet_name
        })
    
    def add_chart(self,
                  sheet_name: str,
                  chart_type: str,
                  title: str,
                  data_range: str,
                  categories_range: Optional[str] = None,
                  position: str = "E2",
                  data_sheet: Optional[str] = None):
        """
        Add chart to sheet
        
        Args:
            sheet_name: Sheet to add chart to
            chart_type: 'bar', 'line', 'pie', 'scatter'
            title: Chart title
            data_range: Range with data (e.g., "B2:B10")
            categories_range: Range with categories (e.g., "A2:A10")
            position: Where to place chart (e.g., "E2")
            data_sheet: Sheet name for data (if different)
        """
        ws = self.sheets.get(sheet_name, self.ws)
        
        # Get data sheet
        if data_sheet:
            data_ws = self.sheets.get(data_sheet, self.ws)
        else:
            data_ws = ws
        
        # Create chart based on type
        if chart_type.lower() == 'bar':
            chart = BarChart()
        elif chart_type.lower() == 'line':
            chart = LineChart()
        elif chart_type.lower() == 'pie':
            chart = PieChart()
        elif chart_type.lower() == 'scatter':
            chart = ScatterChart()
        else:
            chart = BarChart()  # Default
        
        chart.title = title
        chart.style = 10
        
        # Add data
        if data_sheet:
            data_ref = f"'{data_sheet}'!{data_range}"
            cat_ref = f"'{data_sheet}'!{categories_range}" if categories_range else None
        else:
            data_ref = data_range
            cat_ref = categories_range
        
        chart.add_data(data_ws[data_range], titles_from_data=True)
        
        if cat_ref:
            chart.set_categories(data_ws[categories_range])
        
        # Position chart
        ws.add_chart(chart, position)
        
        self.charts.append({
            'type': chart_type,
            'title': title,
            'sheet': sheet_name,
            'position': position
        })
        
        return chart
    
    def style_cell(self, sheet_name: str, cell: str, 
                   font_bold: bool = False,
                   font_color: str = "000000",
                   fill_color: Optional[str] = None,
                   align: str = "left"):
        """Style a cell"""
        ws = self.sheets.get(sheet_name, self.ws)
        cell_obj = ws[cell]
        
        cell_obj.font = Font(bold=font_bold, color=font_color)
        
        if fill_color:
            cell_obj.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        if align == "center":
            cell_obj.alignment = Alignment(horizontal="center")
        elif align == "right":
            cell_obj.alignment = Alignment(horizontal="right")
    
    def auto_size_columns(self, sheet_name: str):
        """Auto-size columns"""
        ws = self.sheets.get(sheet_name, self.ws)
        
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def save(self, filepath: str):
        """Save workbook to file"""
        # Auto-size all sheets
        for sheet_name in self.sheets:
            self.auto_size_columns(sheet_name)
        
        self.wb.save(filepath)
    
    def get_summary(self) -> Dict[str, Any]:
        """Get summary of created Excel file"""
        return {
            'sheets': list(self.sheets.keys()),
            'formulas_count': len(self.formulas),
            'charts_count': len(self.charts),
            'formulas': self.formulas,
            'charts': self.charts
        }


def build_excel_from_spec(spec: Dict[str, Any]) -> bytes:
    """
    Build Excel file from specification
    
    Spec format:
    {
        "sheets": [
            {
                "name": "Sales",
                "data": [["Product", "Qty", "Price"], ["A", 10, 100], ...],
                "formulas": [
                    {"type": "VLOOKUP", "cell": "D2", "lookup_value": "A2", 
                     "table_array": "F1:H10", "col_index_num": 2},
                    ...
                ],
                "charts": [
                    {"type": "bar", "title": "Sales Chart", "data_range": "B2:B10",
                     "categories_range": "A2:A10", "position": "E2"}
                ]
            }
        ]
    }
    """
    builder = ExcelBuilderService()
    
    for sheet_spec in spec.get('sheets', []):
        sheet_name = sheet_spec.get('name', 'Sheet1')
        ws = builder.create_sheet(sheet_name, make_active=True)
        
        # Add data
        if 'data' in sheet_spec:
            builder.add_data(sheet_name, sheet_spec['data'])
        
        # Add formulas
        for formula_spec in sheet_spec.get('formulas', []):
            formula_type = formula_spec.get('type', '').upper()
            
            if formula_type == 'VLOOKUP':
                builder.add_vlookup(
                    sheet_name=sheet_name,
                    target_cell=formula_spec['cell'],
                    lookup_value=formula_spec['lookup_value'],
                    table_array=formula_spec['table_array'],
                    col_index_num=formula_spec['col_index_num'],
                    range_lookup=formula_spec.get('range_lookup', False),
                    lookup_sheet=formula_spec.get('lookup_sheet')
                )
            elif formula_type == 'HLOOKUP':
                builder.add_hlookup(
                    sheet_name=sheet_name,
                    target_cell=formula_spec['cell'],
                    lookup_value=formula_spec['lookup_value'],
                    table_array=formula_spec['table_array'],
                    row_index_num=formula_spec['row_index_num'],
                    range_lookup=formula_spec.get('range_lookup', False),
                    lookup_sheet=formula_spec.get('lookup_sheet')
                )
            else:
                # Custom formula
                builder.add_formula(
                    sheet_name=sheet_name,
                    cell=formula_spec['cell'],
                    formula=formula_spec.get('formula', '')
                )
        
        # Add charts
        for chart_spec in sheet_spec.get('charts', []):
            builder.add_chart(
                sheet_name=sheet_name,
                chart_type=chart_spec.get('type', 'bar'),
                title=chart_spec.get('title', 'Chart'),
                data_range=chart_spec['data_range'],
                categories_range=chart_spec.get('categories_range'),
                position=chart_spec.get('position', 'E2'),
                data_sheet=chart_spec.get('data_sheet')
            )
    
    # Save to bytes
    import io
    output = io.BytesIO()
    builder.save(output)
    output.seek(0)
    
    return output.getvalue()
