"""
File Analyzer Service
AI-powered analysis of Excel files to understand structure, data, and insights
"""
import pandas as pd
import openpyxl
import xlrd
import io
import json
import logging
from typing import Dict, Any, Optional, List, BinaryIO
import re

from app.services.ai_service import invoke_llm

logger = logging.getLogger(__name__)


class FileAnalyzerService:
    """Service to analyze Excel files and generate insights"""

    def __init__(self):
        self.max_sample_rows = 1000  # Analyze first 1000 rows for performance

    async def analyze_excel_file(
        self,
        file_content: BinaryIO,
        filename: str,
        max_rows: int = 1000
    ) -> Dict[str, Any]:
        """
        Analyze Excel file and generate comprehensive insights

        Args:
            file_content: Excel file binary data
            filename: Original filename
            max_rows: Maximum rows to analyze (for performance)

        Returns:
            dict: Analysis results with insights, structure, and recommendations
        """
        try:
            # Read Excel file
            file_bytes = file_content.read() if hasattr(file_content, 'read') else file_content
            file_ext = filename.lower().split('.')[-1]

            if file_ext == 'xlsx':
                workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
                sheets_data = self._parse_xlsx(workbook, max_rows)
            elif file_ext == 'xls':
                workbook = xlrd.open_workbook(file_contents=file_bytes)
                sheets_data = self._parse_xls(workbook, max_rows)
            elif file_ext == 'csv':
                sheets_data = self._parse_csv(file_bytes, filename, max_rows)
            else:
                raise ValueError(f"Unsupported file format: {file_ext}")

            # Analyze each sheet
            analysis_results = []
            for sheet_data in sheets_data:
                analysis = await self._analyze_sheet(sheet_data, filename)
                analysis_results.append(analysis)

            # Generate overall summary
            overall_summary = await self._generate_overall_summary(analysis_results, filename)

            return {
                "filename": filename,
                "file_type": file_ext.upper(),
                "sheet_count": len(sheets_data),
                "sheets": analysis_results,
                "overall_summary": overall_summary,
                "recommendations": self._generate_recommendations(analysis_results)
            }

        except Exception as e:
            logger.error(f"Error analyzing file: {str(e)}")
            raise Exception(f"File analysis failed: {str(e)}")

    def _parse_xlsx(self, workbook, max_rows: int) -> List[Dict]:
        """Parse .xlsx file"""
        sheets_data = []

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            data = []

            # Read rows (limit for performance)
            for idx, row in enumerate(worksheet.iter_rows(values_only=True), 1):
                if idx > max_rows + 1:  # +1 for header
                    break
                if idx == 1:
                    headers = [str(cell) if cell is not None else f"Column{i+1}"
                              for i, cell in enumerate(row)]
                else:
                    data.append([cell if cell is not None else "" for cell in row])

            sheets_data.append({
                'name': sheet_name,
                'headers': headers,
                'rows': data[:max_rows]
            })

        return sheets_data

    def _parse_xls(self, workbook, max_rows: int) -> List[Dict]:
        """Parse .xls file"""
        sheets_data = []

        for sheet_name in workbook.sheet_names():
            sheet = workbook.sheet_by_name(sheet_name)

            headers = [str(sheet.cell_value(0, col)) for col in range(sheet.ncols)]
            rows = []

            for row_idx in range(1, min(sheet.nrows, max_rows + 1)):
                row = [sheet.cell_value(row_idx, col) for col in range(sheet.ncols)]
                rows.append(row)

            sheets_data.append({
                'name': sheet_name,
                'headers': headers,
                'rows': rows
            })

        return sheets_data

    def _parse_csv(self, file_bytes: bytes, filename: str, max_rows: int) -> List[Dict]:
        """Parse CSV file"""
        for encoding in ['utf-8', 'latin-1', 'cp1252']:
            try:
                df = pd.read_csv(io.BytesIO(file_bytes), encoding=encoding, nrows=max_rows)
                break
            except UnicodeDecodeError:
                continue
        else:
            raise ValueError("Could not decode CSV file")

        # Convert DataFrame to list of lists safely
        rows = df.values.tolist() if len(df) > 0 else []
        
        return [{
            'name': filename.replace('.csv', ''),
            'headers': list(df.columns),
            'rows': rows
        }]

    async def _analyze_sheet(self, sheet_data: Dict, filename: str) -> Dict[str, Any]:
        """Analyze a single sheet and generate insights"""

        # Basic statistics
        row_count = len(sheet_data['rows'])
        col_count = len(sheet_data['headers'])

        # Create DataFrame for analysis
        try:
            df = pd.DataFrame(sheet_data['rows'], columns=sheet_data['headers'])
        except Exception as e:
            logger.warning(f"Error creating DataFrame: {str(e)}")
            return {
                'name': sheet_data['name'],
                'row_count': row_count,
                'column_count': col_count,
                'error': str(e)
            }

        # Column analysis
        column_analysis = []
        numeric_columns = []
        categorical_columns = []
        date_columns = []
        text_columns = []

        for col in df.columns:
            try:
                # Get column as Series - ensure it's a Series, not DataFrame
                col_data = df[col]
                if isinstance(col_data, pd.DataFrame):
                    # If somehow we got a DataFrame, take first column
                    col_data = col_data.iloc[:, 0]
                
                # Get sample values safely
                col_series = col_data.dropna().head(5)
                if isinstance(col_series, pd.Series):
                    sample_values = col_series.tolist()
                elif hasattr(col_series, 'values'):
                    sample_values = col_series.values.tolist() if len(col_series) > 0 else []
                else:
                    sample_values = list(col_series) if len(col_series) > 0 else []
            except Exception as e:
                logger.warning(f"Error getting sample values for column {col}: {str(e)}")
                sample_values = []
            
            col_info = {
                'name': str(col),  # Ensure name is string
                'type': 'unknown',
                'null_count': int(df[col].isna().sum()),
                'null_percentage': float((df[col].isna().sum() / len(df)) * 100) if len(df) > 0 else 0.0,
                'unique_count': int(df[col].nunique()),
                'sample_values': sample_values
            }

            # Determine type
            try:
                numeric_data = pd.to_numeric(df[col], errors='coerce')
                if numeric_data.notna().sum() / len(df) > 0.5:
                    col_info['type'] = 'numeric'
                    col_info['min'] = float(numeric_data.min())
                    col_info['max'] = float(numeric_data.max())
                    col_info['mean'] = float(numeric_data.mean())
                    col_info['median'] = float(numeric_data.median())
                    numeric_columns.append(col)
                else:
                    col_info['type'] = 'text'
                    text_columns.append(col)
            except:
                col_info['type'] = 'text'
                text_columns.append(col)

            # Check if categorical
            unique_ratio = df[col].nunique() / len(df)
            if unique_ratio < 0.3 and col_info['type'] != 'numeric':
                col_info['type'] = 'categorical'
                categorical_columns.append(col)

            # Check if date
            try:
                pd.to_datetime(df[col], errors='raise')
                col_info['type'] = 'date'
                date_columns.append(col)
            except:
                pass

            column_analysis.append(col_info)

        # Data quality issues
        quality_issues = []
        for col_info in column_analysis:
            if col_info['null_percentage'] > 20:
                quality_issues.append({
                    'type': 'missing_data',
                    'column': col_info['name'],
                    'severity': 'high' if col_info['null_percentage'] > 50 else 'medium',
                    'message': f"{col_info['name']} has {col_info['null_percentage']:.1f}% missing values"
                })

        # Duplicate rows
        duplicate_count = df.duplicated().sum()
        if duplicate_count > 0:
            quality_issues.append({
                'type': 'duplicates',
                'count': int(duplicate_count),
                'severity': 'medium',
                'message': f"{duplicate_count} duplicate rows found"
            })

        # Generate AI summary
        ai_summary = await self._generate_ai_summary(sheet_data, column_analysis, df)

        return {
            'name': sheet_data['name'],
            'row_count': row_count,
            'column_count': col_count,
            'columns': column_analysis,
            'numeric_columns': numeric_columns,
            'categorical_columns': categorical_columns,
            'date_columns': date_columns,
            'text_columns': text_columns,
            'quality_issues': quality_issues,
            'duplicate_rows': int(duplicate_count),
            'ai_summary': ai_summary,
            'data_preview': df.head(10).to_dict('records') if len(df) > 0 else []
        }

    async def _generate_ai_summary(
        self,
        sheet_data: Dict,
        column_analysis: List[Dict],
        df: pd.DataFrame
    ) -> Dict[str, Any]:
        """Generate AI-powered summary of the data"""

        # Prepare summary for AI
        summary_text = f"""
        Excel Sheet Analysis:
        - Sheet Name: {sheet_data['name']}
        - Rows: {len(sheet_data['rows'])}
        - Columns: {len(sheet_data['headers'])}
        
        Columns:
        {json.dumps([{'name': c['name'], 'type': c['type']} for c in column_analysis], indent=2)}
        
        Sample Data (first 5 rows):
        {df.head(5).to_string() if len(df) > 0 else 'No data'}
        """

        prompt = f"""
        Analyze this Excel file data and provide:
        1. What type of data this appears to be (sales, financial, inventory, etc.)
        2. Key patterns or trends visible
        3. Data quality assessment
        4. Potential use cases
        5. Suggested operations (cleaning, analysis, visualization)
        
        {summary_text}
        
        Respond with JSON:
        {{
            "data_type": "description of data type",
            "key_insights": ["insight1", "insight2", "insight3"],
            "data_quality": "good|fair|poor",
            "use_cases": ["use case 1", "use case 2"],
            "suggested_operations": ["operation1", "operation2"],
            "summary": "Brief 2-3 sentence summary"
        }}
        """

        try:
            response = await invoke_llm(
                prompt=prompt,
                response_schema={"type": "json_object"},
                max_tokens=1500
            )
            return response
        except Exception as e:
            logger.warning(f"Error generating AI summary: {str(e)}")
            return {
                "data_type": "Unknown",
                "key_insights": [],
                "data_quality": "unknown",
                "use_cases": [],
                "suggested_operations": [],
                "summary": "Unable to generate AI summary"
            }

    async def _generate_overall_summary(
        self,
        analysis_results: List[Dict],
        filename: str
    ) -> Dict[str, Any]:
        """Generate overall summary for multi-sheet files"""

        total_rows = sum(sheet['row_count'] for sheet in analysis_results)
        total_columns = sum(sheet['column_count'] for sheet in analysis_results)
        all_quality_issues = []
        for sheet in analysis_results:
            all_quality_issues.extend(sheet.get('quality_issues', []))

        return {
            "total_sheets": len(analysis_results),
            "total_rows": total_rows,
            "total_columns": total_columns,
            "total_quality_issues": len(all_quality_issues),
            "file_size_category": self._categorize_file_size(total_rows),
            "complexity": self._assess_complexity(analysis_results)
        }

    def _categorize_file_size(self, row_count: int) -> str:
        """Categorize file by size"""
        if row_count < 100:
            return "small"
        elif row_count < 1000:
            return "medium"
        elif row_count < 10000:
            return "large"
        else:
            return "very_large"

    def _assess_complexity(self, analysis_results: List[Dict]) -> str:
        """Assess overall complexity"""
        total_columns = sum(sheet['column_count'] for sheet in analysis_results)
        total_issues = sum(len(sheet.get('quality_issues', [])) for sheet in analysis_results)

        if total_columns > 20 or total_issues > 10:
            return "high"
        elif total_columns > 10 or total_issues > 5:
            return "medium"
        else:
            return "low"

    def _generate_recommendations(self, analysis_results: List[Dict]) -> List[str]:
        """Generate actionable recommendations"""
        recommendations = []

        for sheet in analysis_results:
            # Missing data recommendations
            high_missing = [c for c in sheet.get('columns', [])
                           if c.get('null_percentage', 0) > 50]
            if high_missing:
                recommendations.append(
                    f"Sheet '{sheet['name']}': Consider removing or filling columns with >50% missing data"
                )

            # Duplicate recommendations
            if sheet.get('duplicate_rows', 0) > 0:
                recommendations.append(
                    f"Sheet '{sheet['name']}': Remove {sheet['duplicate_rows']} duplicate rows"
                )

            # Type recommendations
            if len(sheet.get('numeric_columns', [])) > 0 and len(sheet.get('categorical_columns', [])) > 0:
                recommendations.append(
                    f"Sheet '{sheet['name']}': Good candidate for charts and visualizations"
                )

        if not recommendations:
            recommendations.append("Data looks clean! Ready for analysis.")

        return recommendations
