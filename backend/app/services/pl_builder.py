"""
P&L (Profit & Loss) Builder Service
Generates Excel files with formulas, charts, and formatting from natural language
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
import pandas as pd
import io
import json
import logging
from typing import Dict, Any, Optional, List
from datetime import datetime, timedelta
import re

from app.services.ai_service import invoke_llm

logger = logging.getLogger(__name__)


class PLBuilderService:
    """Service to generate P&L Excel files from natural language"""

    def __init__(self):
        self.theme_colors = {
            'header': '366092',  # Dark blue
            'revenue': '70AD47',  # Green
            'expense': 'C00000',  # Red
            'net': 'FFC000',  # Orange
            'border': '000000',  # Black
            'light_bg': 'F2F2F2',  # Light gray
        }

    async def generate_pl_from_natural_language(
        self,
        prompt: str,
        user_context: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """
        Generate P&L Excel file from natural language description

        Args:
            prompt: Natural language description (e.g., "Create monthly P&L for 2024")
            user_context: Additional context (company name, currency, etc.)

        Returns:
            bytes: Excel file data
        """
        try:
            # Parse natural language using AI
            pl_spec = await self._parse_pl_request(prompt, user_context)

            # Generate Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Profit & Loss"

            # Build P&L structure
            self._build_pl_structure(ws, pl_spec)

            # Add formulas
            self._add_formulas(ws, pl_spec)

            # Add charts
            self._add_charts(ws, pl_spec)

            # Format worksheet
            self._format_worksheet(ws, pl_spec)

            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            return output.read()

        except Exception as e:
            logger.error(f"Error generating P&L: {str(e)}")
            raise Exception(f"P&L generation failed: {str(e)}")

    async def _parse_pl_request(
        self,
        prompt: str,
        user_context: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """Parse natural language request into structured P&L specification"""

        system_prompt = """
        You are a financial statement generator. Parse the user's request and generate a JSON specification for a Profit & Loss statement.

        Respond with JSON containing:
        {
            "period_type": "monthly|quarterly|yearly",
            "period_count": 12,
            "start_date": "2024-01-01",
            "revenue_categories": ["Sales", "Services", "Other Income"],
            "expense_categories": ["Cost of Goods Sold", "Operating Expenses", "Marketing", "R&D", "Administrative"],
            "currency": "USD",
            "company_name": "Company Name",
            "include_charts": true,
            "include_percentages": true,
            "notes": "Additional notes"
        }

        Extract as much information as possible from the prompt. Use defaults for missing information.
        """

        full_prompt = f"""
        {prompt}
        
        {f'Additional context: {json.dumps(user_context)}' if user_context else ''}
        """

        try:
            response = await invoke_llm(
                prompt=full_prompt,
                response_schema={"type": "json_object"},
                max_tokens=2000
            )

            # Validate and set defaults
            spec = {
                "period_type": response.get("period_type", "monthly"),
                "period_count": response.get("period_count", 12),
                "start_date": response.get("start_date", datetime.now().strftime("%Y-01-01")),
                "revenue_categories": response.get("revenue_categories", ["Revenue", "Other Income"]),
                "expense_categories": response.get("expense_categories", ["Cost of Goods Sold", "Operating Expenses"]),
                "currency": response.get("currency", "USD"),
                "company_name": response.get("company_name", "Company"),
                "include_charts": response.get("include_charts", True),
                "include_percentages": response.get("include_percentages", True),
                "notes": response.get("notes", "")
            }

            return spec

        except Exception as e:
            logger.error(f"Error parsing PL request: {str(e)}")
            # Return default spec if parsing fails
            return {
                "period_type": "monthly",
                "period_count": 12,
                "start_date": datetime.now().strftime("%Y-01-01"),
                "revenue_categories": ["Revenue", "Other Income"],
                "expense_categories": ["Cost of Goods Sold", "Operating Expenses"],
                "currency": "USD",
                "company_name": "Company",
                "include_charts": True,
                "include_percentages": True,
                "notes": ""
            }

    def _build_pl_structure(self, ws, spec: Dict[str, Any]):
        """Build the P&L structure in Excel"""

        row = 1

        # Title
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = f"{spec['company_name']} - Profit & Loss Statement"
        cell.font = Font(size=16, bold=True)
        row += 2

        # Period header
        periods = self._generate_periods(spec)
        header_row = row

        # Column headers
        ws[f'A{header_row}'] = "Category"
        ws[f'B{header_row}'] = "Type"
        for idx, period in enumerate(periods, start=3):
            col_letter = get_column_letter(idx)
            ws[f'{col_letter}{header_row}'] = period

        # Total column
        total_col = len(periods) + 3
        total_col_letter = get_column_letter(total_col)
        ws[f'{total_col_letter}{header_row}'] = "Total"

        row += 1

        # Revenue section
        ws[f'A{row}'] = "REVENUE"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        revenue_start_row = row
        for category in spec['revenue_categories']:
            ws[f'A{row}'] = category
            ws[f'B{row}'] = "Revenue"
            # Add sample data (in real app, these would be user inputs)
            for idx in range(3, total_col):
                col_letter = get_column_letter(idx)
                ws[f'{col_letter}{row}'] = 0  # Placeholder - user fills
            row += 1

        revenue_end_row = row - 1

        # Total Revenue
        ws[f'A{row}'] = "Total Revenue"
        ws[f'A{row}'].font = Font(bold=True)
        for idx in range(3, total_col + 1):
            col_letter = get_column_letter(idx)
            ws[f'{col_letter}{row}'] = f"=SUM({get_column_letter(idx)}{revenue_start_row}:{get_column_letter(idx)}{revenue_end_row})"
        row += 2

        # Expense section
        ws[f'A{row}'] = "EXPENSES"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        expense_start_row = row
        for category in spec['expense_categories']:
            ws[f'A{row}'] = category
            ws[f'B{row}'] = "Expense"
            for idx in range(3, total_col):
                col_letter = get_column_letter(idx)
                ws[f'{col_letter}{row}'] = 0  # Placeholder
            row += 1

        expense_end_row = row - 1

        # Total Expenses
        ws[f'A{row}'] = "Total Expenses"
        ws[f'A{row}'].font = Font(bold=True)
        for idx in range(3, total_col + 1):
            col_letter = get_column_letter(idx)
            ws[f'{col_letter}{row}'] = f"=SUM({get_column_letter(idx)}{expense_start_row}:{get_column_letter(idx)}{expense_end_row})"
        row += 2

        # Net Profit/Loss
        revenue_total_col = get_column_letter(total_col)
        expense_total_col = get_column_letter(total_col)
        revenue_row = revenue_end_row + 1
        expense_row = expense_end_row + 1

        ws[f'A{row}'] = "NET PROFIT / (LOSS)"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        for idx in range(3, total_col + 1):
            col_letter = get_column_letter(idx)
            ws[f'{col_letter}{row}'] = f"={get_column_letter(idx)}{revenue_row}-{get_column_letter(idx)}{expense_row}"

        # Store structure info for formulas and charts
        spec['structure'] = {
            'header_row': header_row,
            'revenue_start': revenue_start_row,
            'revenue_end': revenue_end_row,
            'revenue_total_row': revenue_row,
            'expense_start': expense_start_row,
            'expense_end': expense_end_row,
            'expense_total_row': expense_row,
            'net_row': row,
            'total_col': total_col,
            'periods': periods
        }

    def _generate_periods(self, spec: Dict[str, Any]) -> List[str]:
        """Generate period labels based on specification"""
        periods = []
        start_date = datetime.strptime(spec['start_date'], "%Y-%m-%d")
        period_type = spec['period_type']
        count = spec['period_count']

        for i in range(count):
            if period_type == "monthly":
                date = start_date + timedelta(days=30 * i)
                periods.append(date.strftime("%b %Y"))
            elif period_type == "quarterly":
                date = start_date + timedelta(days=90 * i)
                quarter = (date.month - 1) // 3 + 1
                periods.append(f"Q{quarter} {date.year}")
            else:  # yearly
                date = start_date + timedelta(days=365 * i)
                periods.append(str(date.year))

        return periods

    def _add_formulas(self, ws, spec: Dict[str, Any]):
        """Add Excel formulas for calculations"""
        # Formulas are already added in _build_pl_structure
        # This method can be extended for additional formulas
        pass

    def _add_charts(self, ws, spec: Dict[str, Any]):
        """Add charts to the worksheet"""
        if not spec.get('include_charts', True):
            return

        structure = spec['structure']
        chart_row = structure['net_row'] + 3

        # Revenue vs Expenses Chart
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Revenue vs Expenses"
        chart.y_axis.title = "Amount"
        chart.x_axis.title = "Period"

        # Data ranges
        periods = structure['periods']
        data_cols = [get_column_letter(i) for i in range(3, 3 + len(periods))]

        # Revenue data
        revenue_data = Reference(
            ws,
            min_col=3,
            min_row=structure['revenue_total_row'],
            max_col=3 + len(periods) - 1,
            max_row=structure['revenue_total_row']
        )

        # Expense data
        expense_data = Reference(
            ws,
            min_col=3,
            min_row=structure['expense_total_row'],
            max_col=3 + len(periods) - 1,
            max_row=structure['expense_total_row']
        )

        # Categories
        cats = Reference(
            ws,
            min_col=3,
            min_row=structure['header_row'],
            max_col=3 + len(periods) - 1,
            max_row=structure['header_row']
        )

        chart.add_data(revenue_data, titles_from_data=True)
        chart.add_data(expense_data, titles_from_data=True)
        chart.set_categories(cats)

        # Position chart
        ws.add_chart(chart, f"A{chart_row}")

    def _format_worksheet(self, ws, spec: Dict[str, Any]):
        """Apply formatting to the worksheet"""

        # Header row formatting
        header_row = spec['structure']['header_row']
        for cell in ws[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.theme_colors['header'], end_color=self.theme_colors['header'], fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        for idx in range(3, spec['structure']['total_col'] + 1):
            col_letter = get_column_letter(idx)
            ws.column_dimensions[col_letter].width = 15

        # Number formatting
        for row in ws.iter_rows(min_row=header_row + 1, max_row=spec['structure']['net_row']):
            for cell in row[2:]:  # Skip first two columns
                if cell.value and isinstance(cell.value, (int, float)):
                    cell.number_format = f'$#,##0.00'

        # Revenue rows - green
        for row in range(spec['structure']['revenue_start'], spec['structure']['revenue_end'] + 1):
            for cell in ws[row]:
                if cell.column > 2:  # Data columns
                    cell.fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type="solid")

        # Expense rows - light red
        for row in range(spec['structure']['expense_start'], spec['structure']['expense_end'] + 1):
            for cell in ws[row]:
                if cell.column > 2:
                    cell.fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type="solid")

        # Net row - highlight
        net_row = spec['structure']['net_row']
        for cell in ws[net_row]:
            cell.font = Font(bold=True, size=12)
            if cell.column > 2:
                cell.fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type="solid")
