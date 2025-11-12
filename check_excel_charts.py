#!/usr/bin/env python3
"""
Check what openpyxl sees in the Excel file
"""
import openpyxl

filename = "Ch_01_ChartEssentials_test.xlsx"

print("="*70)
print(f"Analyzing Excel File: {filename}")
print("="*70)

try:
    # Load workbook
    wb = openpyxl.load_workbook(filename, data_only=True)
    print(f"\n✅ File loaded successfully")
    print(f"   Sheets: {wb.sheetnames}")

    # Check each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n📊 Sheet: {sheet_name}")

        # Check dimensions
        print(f"   Dimensions: {ws.dimensions}")
        print(f"   Max row: {ws.max_row}")
        print(f"   Max column: {ws.max_column}")

        # Check for charts (openpyxl can see these)
        if hasattr(ws, '_charts'):
            charts = ws._charts
            print(f"   Charts (openpyxl): {len(charts)}")
            for i, chart in enumerate(charts, 1):
                print(f"      Chart {i}: {type(chart).__name__}")
        else:
            print(f"   Charts (openpyxl): Not accessible via openpyxl")

        # Try to get chart objects
        try:
            from openpyxl.chart import *
            chart_count = 0
            for obj in dir(ws):
                if 'chart' in obj.lower():
                    print(f"   Found chart-related: {obj}")
        except:
            pass

        # Show some data
        print(f"\n   Sample data (first 5 rows):")
        for row_idx, row in enumerate(ws.iter_rows(max_row=5, values_only=True), 1):
            row_data = [str(cell)[:15] if cell is not None else "" for cell in row]
            print(f"      Row {row_idx}: {row_data[:5]}")  # First 5 columns

    print("\n" + "="*70)
    print("IMPORTANT NOTE:")
    print("="*70)
    print("❌ openpyxl CANNOT export existing Excel charts as images!")
    print("❌ openpyxl can only CREATE new charts programmatically!")
    print("\nTo export actual Excel charts, you need:")
    print("  ✅ Windows OS")
    print("  ✅ Microsoft Excel installed")
    print("  ✅ pywin32 package")
    print("  ✅ Windows COM automation (excel_to_ppt_windows.py)")
    print("="*70)

except Exception as e:
    print(f"\n❌ Error: {e}")
    import traceback
    traceback.print_exc()
